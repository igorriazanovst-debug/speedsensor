"""
Модель данных эксперимента и экспорт.
"""
import csv
import math
from dataclasses import dataclass, field, asdict
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


@dataclass
class DataRow:
    timestamp_s:  float
    omega_rad_s:  float

    @property
    def rps(self) -> float:
        return self.omega_rad_s / (2.0 * math.pi)

    @property
    def rpm(self) -> float:
        return self.rps * 60.0

    def linear_speed_mm_s(self, disk_diameter_mm: float) -> float:
        """Линейная скорость точки на ободе диска, мм/с."""
        r = disk_diameter_mm / 2.0
        return self.omega_rad_s * r


class ExperimentData:
    HEADERS = ["Время, с", "ω, рад/с", "ω, об/с", "RPM", "V, мм/с"]

    def __init__(self, disk_diameter_mm: float = 75.0):
        self.disk_diameter_mm = disk_diameter_mm
        self._rows: list[DataRow] = []

    # ---------------------------------------------------------------- data --

    def append(self, row: DataRow):
        self._rows.append(row)

    def clear(self):
        self._rows.clear()

    def __len__(self):
        return len(self._rows)

    def __getitem__(self, idx):
        return self._rows[idx]

    def row_as_display(self, idx: int) -> list[str]:
        r = self._rows[idx]
        return [
            f"{r.timestamp_s:.3f}",
            f"{r.omega_rad_s:.4f}",
            f"{r.rps:.4f}",
            f"{r.rpm:.2f}",
            f"{r.linear_speed_mm_s(self.disk_diameter_mm):.2f}",
        ]

    def all_as_lists(self) -> list[list]:
        return [self.row_as_display(i) for i in range(len(self._rows))]

    # -------------------------------------------------------------- export --

    def export_csv(self, path: str, delimiter: str = ";") -> None:
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f, delimiter=delimiter)
            writer.writerow(self.HEADERS)
            for i in range(len(self._rows)):
                writer.writerow(self.row_as_display(i))

    def export_xlsx(self, path: str) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Данные эксперимента"

        # Заголовок
        header_fill = PatternFill("solid", fgColor="1E1E2E")
        header_font = Font(bold=True, color="CBA6F7", name="Calibri", size=11)
        for col, h in enumerate(self.HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Данные
        alt_fill = PatternFill("solid", fgColor="181825")
        norm_fill = PatternFill("solid", fgColor="1E1E2E")
        data_font = Font(name="Calibri", size=10, color="CDD6F4")

        for row_idx, i in enumerate(range(len(self._rows)), 2):
            row_data = self.row_as_display(i - 2)
            fill = alt_fill if row_idx % 2 == 0 else norm_fill
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.fill = fill
                cell.font = data_font
                cell.alignment = Alignment(horizontal="right")

        # Ширина столбцов
        col_widths = [14, 14, 14, 12, 14]
        for col, w in enumerate(col_widths, 1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col)
            ].width = w

        # Метаданные
        ws2 = wb.create_sheet("Метаданные")
        ws2["A1"] = "Дата экспорта"
        ws2["B1"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws2["A2"] = "Диаметр диска, мм"
        ws2["B2"] = self.disk_diameter_mm
        ws2["A3"] = "Записей"
        ws2["B3"] = len(self._rows)

        wb.save(path)
